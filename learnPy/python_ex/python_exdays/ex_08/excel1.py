#!usr/bin/python3
# -*- coding:utf-8 -*-

"""
创建excel文件

version：1.0
date：2019\6\15
"""

from openpyxl import Workbook
from openpyxl.worksheet.table import Table,TableStyleInfo

workbook = Workbook()
sheet = workbook.active
data = [
    ["1001",'白远方','男','13123456789'],
    ["1002","白洁","女","1323345566"]
    ]
sheet.append(["学号","姓名","性别","电话"])
for row in data:
    sheet.append(row)
tab = Table(displayName="Table1",ref="A1:D4")
#当ref="A1:E5"时，会报column headings must be strings，因为D的列标题空

tab.tableStyleInfo = TableStyleInfo(
    name = "TableStyleMedium9",showFirstColumn=False,
    showLastColumn=False,showRowStripes=True,showColumnStripes=True
    )
sheet.add_table(tab)
workbook.save("./res/全班学生数据.xlsx")


"""
读取Excel文件
"""
from openpyxl import load_workbook

workbook = load_workbook("./res/全班学生数据.xlsx")
print(workbook.sheetnames)
sheet=workbook[workbook.sheetnames[0]]
print(sheet.title)
for row in range(2,7):
    for col in range(65,70):
        cell_index = chr(col) + str(row)
        print(sheet[cell_index].value,end='\t')
    print()
